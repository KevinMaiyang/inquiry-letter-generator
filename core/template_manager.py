# core/template_manager.py
import os
import sys
import shutil
from openpyxl import load_workbook
from core.utils import get_user_template_path, get_default_template_path

def parse_A9(text):
    """解析 A9 中的地址和联系人"""
    if "回函地址：" in text and "联系人：" in text:
        parts = text.split("联系人：", 1)
        contact = parts[1].strip() if len(parts) > 1 else ""
        addr = parts[0].replace("回函地址：", "").strip()
        return addr, contact
    return text.strip(), ""

class TemplateManager:
    def __init__(self):
        self.user_template = get_user_template_path()
        self._ensure_template_exists()

    def _ensure_template_exists(self):
        """确保用户模板存在，如果不存在则从默认模板复制"""
        if not self.user_template.exists():
            default_template = get_default_template_path()
            if os.path.exists(default_template):
                shutil.copy2(default_template, self.user_template)
                print(f"已创建用户模板: {self.user_template}")
            else:
                # 如果默认模板也不存在，创建一个空的模板文件
                print(f"警告: 默认模板不存在 {default_template}")
                self._create_empty_template()

    def _create_empty_template(self):
        """创建一个空的模板文件"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        # 设置基本的单元格结构
        ws['A9'] = "回函地址：    联系人："
        ws['A10'] = "电话："
        ws['C10'] = "邮箱："
        ws['B19'] = ""
        ws['D20'] = ""
        wb.save(self.user_template)

    def load_fields(self):
        if not self.user_template.exists():
            return self._get_default_fields()
            
        wb = load_workbook(self.user_template, read_only=True, data_only=True)
        ws = wb.active
        a9 = str(ws['A9'].value or "").strip()
        addr, contact = parse_A9(a9)
        return {
            'address': addr,
            'contact': contact,
            'phone': str(ws['A10'].value or "").replace("电话：", "").strip(),
            'email': str(ws['C10'].value or "").replace("邮箱：", "").strip(),
            'issuer': str(ws['B19'].value or "").strip(),
            'date': str(ws['D20'].value or "").strip()
        }

    def _get_default_fields(self):
        """返回默认字段值"""
        return {
            'address': '',
            'contact': '',
            'phone': '',
            'email': '',
            'issuer': '',
            'date': ''
        }

    def save_fields(self, fields):
        wb = load_workbook(self.user_template)
        ws = wb.active
        ws['A9'] = f"回函地址：{fields['address']}    联系人：{fields['contact']}"
        ws['A10'] = f"电话：{fields['phone']}"
        ws['C10'] = f"邮箱：{fields['email']}"
        ws['B19'] = fields['issuer']
        ws['D20'] = fields['date']
        wb.save(self.user_template)