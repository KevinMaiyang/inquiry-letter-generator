# excel_generator.py
from openpyxl import Workbook
from utils import clone_sheet
from openpyxl import load_workbook, Workbook

def generate_excel(data_list, template_path, output_path):
    template_wb = load_workbook(template_path)
    template_ws = template_wb.active

    new_wb = Workbook()
    new_wb.remove(new_wb.active)

    for data in data_list:
        sheet_name = data['sheet_name'][:31]
        ws = clone_sheet(template_ws, new_wb, sheet_name)

        ws['D1'] = f"编号：{data['number']}"
        ws['A3'] = data['unit']
        ws['A4'] = (
            f"    我公司承担的{data['project']}项目，已完成了合同约定的相应工作，我公司核算截止到该项目{data['season']}计价，"
            f"债权记录截止到{data['date']}尚有下表列示数据未收到，请贵公司核对，如与贵单位记录相符，请在本函下端“信息证明无误”处签章证明；"
            f"如有不符，请在“信息不符”处列明不符金额"
        )

        def clean_val(val):
            if val == "" or val is None:
                return "0.00"
            try:
                num = float(str(val).replace(',', ''))
                return f"{num:,.2f}"
            except:
                return "0.00"

        ws['C13'] = clean_val(data['receivable'])
        ws['C14'] = clean_val(data['long_term'])
        ws['C16'] = clean_val(data['total'])

        ws['A9'] = f"回函地址：{data['address']}    联系人：{data['contact']}"
        ws['A10'] = f"电话：{data['phone']}"
        ws['C10'] = f"邮箱：{data['email']}"
        ws['B19'] = data['issuer']
        ws['D20'] = data['date']
        ws['B13'] = data['date']
        ws['B14'] = data['date']
        ws['A29'] = f"3.备注：（如果截止{data['date']}日后情况有变化，请在此列明最新情况）"

    new_wb.save(output_path)
    return output_path