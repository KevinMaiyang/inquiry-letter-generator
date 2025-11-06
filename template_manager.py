# template_manager.py
import os
from openpyxl import load_workbook
from utils import get_user_template_path, resource_path
import shutil

def parse_A9(text):
    """解析 A9 中的地址和联系人"""
    if "回函地址：" in text and "联系人：" in text:
        parts = text.split("联系人：", 1)
        contact = parts[1].strip() if len(parts) > 1 else ""
        addr = parts[0].replace("回函地址：", "").strip()
        return addr, contact
    return text.strip(), ""

class TemplateManager:
    def __init__(self):
        self.user_template = get_user_template_path()
        self._ensure_template_exists()

    def _ensure_template_exists(self):
        if not self.user_template.exists():
            internal = resource_path("template.xlsx")
            if os.path.exists(internal):
                shutil.copy2(internal, self.user_template)

    def load_fields(self):
        wb = load_workbook(self.user_template, read_only=True, data_only=True)
        ws = wb.active
        a9 = str(ws['A9'].value or "").strip()
        addr, contact = parse_A9(a9)
        return {
            'address': addr,
            'contact': contact,
            'phone': str(ws['A10'].value or "").replace("电话：", "").strip(),
            'email': str(ws['C10'].value or "").replace("邮箱：", "").strip(),
            'issuer': str(ws['B19'].value or "").strip(),
            'date': str(ws['D20'].value or "").strip()
        }

    def save_fields(self, fields):
        wb = load_workbook(self.user_template)
        ws = wb.active
        ws['A9'] = f"回函地址：{fields['address']}    联系人：{fields['contact']}"
        ws['A10'] = f"电话：{fields['phone']}"
        ws['C10'] = f"邮箱：{fields['email']}"
        ws['B19'] = fields['issuer']
        ws['D20'] = fields['date']
        wb.save(self.user_template)