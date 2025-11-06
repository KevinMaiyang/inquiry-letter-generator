# utils.py
import sys
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

def resource_path(relative_path):
    """获取资源路径（兼容 PyInstaller）"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def get_user_template_path():
    """获取用户模板路径"""
    if sys.platform == "win32":
        config_dir = Path(os.environ["APPDATA"]) / "InquiryLetterGenerator"
    else:
        config_dir = Path.home() / ".config" / "InquiryLetterGenerator"
    config_dir.mkdir(parents=True, exist_ok=True)
    return config_dir / "template.xlsx"

def parse_date(date_str):
    """解析 'YYYY.M.D' 日期"""
    parts = date_str.split('.')
    if len(parts) != 3:
        raise ValueError("日期格式应为 YYYY.M.D")
    return tuple(map(int, parts))

def get_season_from_date(date_str):
    """计算季度"""
    y, m, d = parse_date(date_str)
    q_ends = [(y, 3, 31), (y, 6, 30), (y, 9, 30), (y, 12, 31)]
    input_date = (y, m, d)
    for i, q_end in enumerate(q_ends):
        if input_date >= q_end:
            quarter = i + 1
        else:
            break
    else:
        quarter = 0
    if quarter == 0:
        return f"{y - 1}年第4季度"
    return f"{y}年第{quarter}季度"

def clone_sheet(source_ws, target_wb, new_title):
    """克隆工作表（含格式）"""
    target_ws = target_wb.create_sheet(title=new_title)
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    if source_ws.merged_cells:
        for merged_range in source_ws.merged_cells:
            target_ws.merge_cells(str(merged_range))

    for col in range(1, source_ws.max_column + 1):
        col_letter = get_column_letter(col)
        if col_letter in source_ws.column_dimensions:
            target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width

    for row in range(1, source_ws.max_row + 1):
        if source_ws.row_dimensions[row].height:
            target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height

    return target_ws